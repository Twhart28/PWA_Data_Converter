import re
import sys
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import pandas as pd
import pdfplumber
from openpyxl.styles import Alignment
from PIL import Image, ImageTk


# Analysis mode selector. Choose 1 to use combined peripheral SYS/DIA/MEAN matching
# or 2 to match only on peripheral systolic pressure.
ANALYSIS_MODE = 2

COLUMNS = [
    "Source File",
    "Patient ID",
    "Scanned ID",
    "Scan Date",
    "Scan Time",
    "Recording #",
    "Analyed",
    "Date of Birth",
    "Age",
    "Gender",
    "Height (m)",
    "# of Pulses",
    "Pulse Height",
    "Pulse Height Variation (%)",
    "Diastolic Variation (%)",
    "Shape Deviation (%)",
    "Pulse Length Variation (%)",
    "Overall Quality (%)",
    "Peripheral Systolic Pressure (mmHg)",
    "Peripheral Diastolic Pressure (mmHg)",
    "Peripheral Pulse Pressure (mmHg)",
    "Peripheral Mean Pressure (mmHg)",
    "Aortic Systolic Pressure (mmHg)",
    "Aortic Diastolic Pressure (mmHg)",
    "Aortic Pulse Pressure (mmHg)",
    "Heart Rate (bpm)",
    "Pulse Pressure Amplification (%)",
    "Period (ms)",
    "Ejection Duration (ms)",
    "Ejection Duration (%)",
    "Aortic T2 (ms)",
    "P1 Height (mmHg)",
    "Aortic Augmentation (mmHg)",
    "Aortic AIx AP/PP(%)",
    "Aortic AIx P2/P1(%)",
    "Aortic AIx AP/PP @ HR75 (%)",
    "Buckberg SEVR (%)",
    "PTI Systolic (mmHg.s/min)",
    "PTI Diastolic (mmHg.s/min)",
    "End Systolic Pressure (mmHg)",
    "MAP Systolic (mmHg)",
    "MAP Diastolic (mmHg)",
]

DETAILED_REPORT_MARKER = "PWA Detailed Report"
CLINICAL_REPORT_MARKER = "PWA Clinical Report"
CLINICAL_REPORT_MESSAGE = (
    "Recognized as a Clinical Report, only upload the Detailed Reports"
)
UNRECOGNIZED_REPORT_MESSAGE = "Not recognized as a PWA Detailed Report"

EXTRA_COLUMNS = ["Source Path"]
AVERAGED_EXCLUDED_FIELDS = {
    "Source File",
    "Scanned ID",
    "Scan Date",
    "Scan Time",
    "Analyed",
    "Recording #",
    "Source Path",
}
CHECKMARK = "✔"
SELECTED_COLOR = "#c8f7c5"
APP_ICON_PATH = Path(__file__).with_name("App_Logo.ico")


def center_window(window: tk.Misc) -> None:
    """Center a Tkinter window on the screen."""

    window.update_idletasks()
    width = window.winfo_width()
    height = window.winfo_height()

    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()

    x_offset = max((screen_width - width) // 2, 0)
    y_offset = max((screen_height - height) // 2, 0)

    window.geometry(f"{width}x{height}+{x_offset}+{y_offset}")


def terminate_application(root: tk.Misc) -> None:
    """Close the Tk root and terminate the entire application."""

    try:
        root.destroy()
    finally:
        sys.exit(0)


def configure_popup_close(window: tk.Toplevel, root: tk.Misc) -> None:
    """Ensure clicking the close button exits the whole application."""

    window.protocol("WM_DELETE_WINDOW", lambda: terminate_application(root))


def set_app_icon(window: tk.Tk | tk.Toplevel) -> None:
    """Apply the application icon to a Tkinter window if available."""

    if APP_ICON_PATH.exists():
        try:
            window.iconbitmap(APP_ICON_PATH)
        except tk.TclError:
            # Fallback gracefully if the platform does not support .ico files
            pass


class LoadingWindow:
    def __init__(self, root: tk.Misc, message: str, total_steps: int | None = None):
        self.window = tk.Toplevel(root)
        self.window.title("PWA Data Analyzer")
        self.window.geometry("340x160")
        self.window.resizable(False, False)
        self.window.grab_set()
        set_app_icon(self.window)
        configure_popup_close(self.window, root)

        self.window.bind("<Unmap>", self._release_grab)
        self.window.bind("<Map>", self._restore_grab)

        self.total_steps = total_steps if total_steps and total_steps > 0 else None

        label = ttk.Label(self.window, text=message, wraplength=300)
        label.pack(pady=(20, 10), padx=10)

        progress_mode = "determinate" if self.total_steps else "indeterminate"
        self.progress = ttk.Progressbar(
            self.window,
            orient="horizontal",
            mode=progress_mode,
            length=280,
            maximum=self.total_steps or 100,
        )
        self.progress.pack(pady=(0, 10))

        if progress_mode == "indeterminate":
            self.progress.start(10)

        self.status_label = ttk.Label(self.window, text="")
        self.status_label.pack(pady=(0, 10))

        center_window(self.window)
        self.window.update()

    def update_progress(self, completed_steps: int) -> None:
        if not self.total_steps:
            return

        completed_steps = min(completed_steps, self.total_steps)
        self.progress["value"] = completed_steps
        self.status_label.config(
            text=f"Processed {completed_steps} of {self.total_steps} files"
        )
        self.window.update_idletasks()

    def _release_grab(self, _event: tk.Event | None = None) -> None:  # type: ignore[type-arg]
        self.window.grab_release()

    def _restore_grab(self, _event: tk.Event | None = None) -> None:  # type: ignore[type-arg]
        self.window.grab_set()

    def close(self) -> None:
        if self.progress.cget("mode") == "indeterminate":
            self.progress.stop()
        self.window.destroy()

README_TEXT = """# PWA Data Converter

**Contact:** thomaswhart28@gmail.com

Convert one or many PWA analysis PDFs into a structured Excel workbook. The app guides you through a short startup dialog, lets you review the README in-app, and walks you through selecting files, choosing the output location, and reviewing multi-entry patients before exporting.

## Prerequisites

Install the dependencies:

```bash
pip install -r requirements.txt
```

## Usage

1. Run the converter:

```bash
python pwa_converter.py
```

2. On startup, use **Read Me** to view this guide in-app or choose **Continue** to proceed.
3. Use the file picker to select one or more PWA PDF reports.
4. Choose where to save the Excel workbook; a timestamped filename is suggested for convenience.
5. Watch the progress dialog as files are analyzed. If a patient has more than two valid entries, you can accept the automatic pairing or launch the **Manual Overview** to pick which records to average.
6. When the export completes, you will see a confirmation with the destination path.

## Key Features

- **In-app README access:** The startup popup offers a **Read Me** button and the window close button behaves the same as clicking **Close** inside the README viewer.
- **PDF validation and processing:** Each PWA Detailed Report is parsed for patient demographics, hemodynamic measurements, waveform quality metrics, and timing values.
- **Automatic and manual record pairing:** Patients with multiple entries are averaged using the configured analysis mode, with an optional manual review step to override the automatic pairing.
- **Organized Excel output:** Three sheets are created—**All Data** (full dataset), **Kept Data** (records included in analysis), and **Averaged Data** (per-patient averages). Dates are normalized, headers and key identifiers are aligned for readability, and analysis flags mark which rows were used.
- **Convenient defaults:** Suggested filenames include timestamps, and progress dialogs keep you informed during analysis and export.

## Extracted Fields

The export preserves a consistent order of fields per PDF, including:

- Source details: Source File and Source Path
- Patient demographics: Patient ID, Date of Birth, Age, Gender, Height (m)
- Signal quality: # of Pulses, Pulse Height, Pulse Height Variation (%), Diastolic Variation (%), Shape Deviation (%), Pulse Length Variation (%), Overall Quality (%)
- Peripheral values: Peripheral Systolic/Diastolic/Mean Pressure (mmHg), Peripheral Pulse Pressure (mmHg)
- Aortic values: Aortic Systolic/Diastolic/Pulse Pressure (mmHg), Aortic AIx (AP/PP, P2/P1, AP/PP @ HR75), Aortic Augmentation (mmHg)
- Cardiac timing: Heart Rate (bpm), Period (ms), Ejection Duration (ms/%), Aortic T2 (ms)
- Additional metrics: Pulse Pressure Amplification (%), P1 Height (mmHg), Buckberg SEVR (%), PTI Systolic/Diastolic (mmHg.s/min), End Systolic Pressure (mmHg), MAP Systolic/Diastolic (mmHg)

The workbook keeps this order across all sheets so downstream analysis remains predictable.
"""

def show_readme_popup(root: tk.Misc) -> None:
    """Display the README contents in a popup window."""

    window = tk.Toplevel(root)
    window.title("Read Me")
    window.geometry("700x500")
    window.resizable(True, True)
    window.grab_set()
    set_app_icon(window)
    window.protocol("WM_DELETE_WINDOW", window.destroy)

    text_frame = ttk.Frame(window)
    text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    text_widget = tk.Text(text_frame, wrap=tk.WORD, yscrollcommand=scrollbar.set)
    text_widget.pack(fill=tk.BOTH, expand=True)
    scrollbar.config(command=text_widget.yview)

    readme_path = Path(__file__).with_name("README.md")
    if readme_path.exists():
        content = readme_path.read_text(encoding="utf-8")
    else:
        content = "README file not found."

    text_widget.insert(tk.END, content)
    text_widget.configure(state="disabled")

    close_button = ttk.Button(window, text="Close", command=window.destroy)
    close_button.pack(pady=10)

    center_window(window)
    root.wait_window(window)

def open_readme(parent: tk.Misc | None = None) -> None:
    """Show a simple readme window with scrollable text."""
    win = tk.Toplevel(parent)
    win.title("Read Me")
    win.resizable(True, True)
    set_app_icon(win)

    # Frame
    frame = ttk.Frame(win)
    frame.pack(fill="both", expand=True, padx=10, pady=10)

    # Text widget + scrollbar
    text = tk.Text(frame, wrap="word", height=25, width=80)
    scroll = ttk.Scrollbar(frame, orient="vertical", command=text.yview)
    text.configure(yscrollcommand=scroll.set)

    text.grid(row=0, column=0, sticky="nsew")
    scroll.grid(row=0, column=1, sticky="ns")

    frame.rowconfigure(0, weight=1)
    frame.columnconfigure(0, weight=1)

    # Insert readme text and make it read-only
    text.insert("1.0", README_TEXT)
    text.configure(state="disabled")

    # Close button
    close_btn = ttk.Button(frame, text="Close", command=win.destroy)
    close_btn.grid(row=1, column=0, columnspan=2, pady=(8, 0))

    # Optional: center over parent if provided
    if parent is not None:
        win.transient(parent)
        win.grab_set()
        
def show_startup_popup(root: tk.Misc) -> bool:
    """Show a startup popup with image, buttons, and version text."""

    window = tk.Toplevel(root)
    window.title("PWA Data Analyzer")
    window.resizable(False, False)
    set_app_icon(window)
    configure_popup_close(window, root)

    proceed = {"continue": False}

    # Main container
    container = ttk.Frame(window)
    container.pack(padx=20, pady=20)

    # --- Image section ---
    # Tries to load an image named 'startup_image.png' in the same folder as this script.
    # You can change this filename/path if you like.
    image_loaded = False
    try:
        possible_names = ["startup_image.png", "statup_image.png"]
        image_path = next(
            (Path(__file__).with_name(name) for name in possible_names if Path(__file__).with_name(name).exists()),
            None,
        )
        if image_path:
            img = Image.open(image_path)
            # Resize gently so it fits on most screens
            img.thumbnail((900, 600), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(img)

            img_label = ttk.Label(container, image=photo)
            img_label.image = photo  # keep reference so it doesn't get GC'd
            img_label.pack()
            image_loaded = True
    except Exception:
        image_loaded = False

    if not image_loaded:
        ttk.Label(
            container,
            text="PWA Data Analyzer",
            font=("TkDefaultFont", 14, "bold"),
        ).pack(pady=(0, 10))

    # --- Buttons row ---
    buttons_frame = ttk.Frame(container)
    buttons_frame.pack(pady=(15, 8))

    read_me_btn = ttk.Button(buttons_frame, text="Read Me", command=lambda: open_readme(window))
    read_me_btn.pack(side=tk.LEFT, padx=10)

    def _continue() -> None:
        proceed["continue"] = True
        window.destroy()

    continue_btn = ttk.Button(buttons_frame, text="Continue", command=_continue)
    continue_btn.pack(side=tk.LEFT, padx=10)

    # --- Version label ---
    version_label = ttk.Label(
        container,
        text="Version 1 (12/07/25)",
        font=("TkDefaultFont", 10),
    )
    version_label.pack(pady=(5, 0))

    # Center and wait until the popup is closed (user can click the X)
    center_window(window)
    root.wait_window(window)
    return proceed["continue"]

def select_input_files(root: tk.Misc | None = None) -> tuple[Path, ...]:
    should_destroy = False
    if root is None:
        root = tk.Tk()
        set_app_icon(root)
        root.withdraw()
        should_destroy = True

    file_paths = filedialog.askopenfilenames(
        title="Select PWA PDF files",
        filetypes=[("PDF Files", "*.pdf")],
        parent=root,
    )
    root.update()

    if should_destroy:
        root.destroy()

    return tuple(Path(path) for path in file_paths)


def select_output_file(root: tk.Misc | None = None) -> Path | None:
    should_destroy = False
    if root is None:
        root = tk.Tk()
        set_app_icon(root)
        root.withdraw()
        should_destroy = True

    timestamp = datetime.now().strftime("%m/%d/%y %H:%M")
    safe_timestamp = timestamp.replace("/", "-").replace(":", "-")
    default_name = f"PWA Export ({safe_timestamp}).xlsx"
    output_path = filedialog.asksaveasfilename(
        title="Save Excel file as",
        initialfile=default_name,
        defaultextension=".xlsx",
        filetypes=[("Excel Workbook", "*.xlsx"), ("All Files", "*.*")],
        parent=root,
    )
    root.update()
    if should_destroy:
        root.destroy()
    if not output_path:
        return None
    return Path(output_path)


def _bind_mousewheel(canvas: tk.Canvas) -> None:
    def _on_mousewheel(event: tk.Event) -> None:  # type: ignore[type-arg]
        if getattr(event, "num", None) == 4 or getattr(event, "delta", 0) > 0:
            canvas.yview_scroll(-1, "units")
        elif getattr(event, "num", None) == 5 or getattr(event, "delta", 0) < 0:
            canvas.yview_scroll(1, "units")

    canvas.bind("<MouseWheel>", _on_mousewheel)
    canvas.bind("<Button-4>", _on_mousewheel)
    canvas.bind("<Button-5>", _on_mousewheel)


def show_pdf_preview(parent: tk.Misc, pdf_path: Path) -> None:
    if not pdf_path.exists():
        messagebox.showerror(
            "PDF Preview",
            f"The file {pdf_path} could not be found.",
            parent=parent,
        )
        return

    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                raise ValueError("PDF has no pages to preview.")

            page_images = [page.to_image(resolution=120).original for page in pdf.pages]
    except Exception as exc:  # noqa: BLE001
        messagebox.showerror(
            "PDF Preview", f"Unable to preview PDF: {exc}", parent=parent
        )
        return

    preview_photos: list[ImageTk.PhotoImage] = []
    for image in page_images:
        preview_image = image.copy()
        preview_image.thumbnail((900, 1200), Image.Resampling.LANCZOS)
        preview_photos.append(ImageTk.PhotoImage(preview_image))

    max_width = max(photo.width() for photo in preview_photos)
    window_width = min(max_width + 40, 1000)
    window_height = 900

    preview_window = tk.Toplevel(parent)
    preview_window.title(pdf_path.name)
    preview_window.protocol("WM_DELETE_WINDOW", preview_window.destroy)
    set_app_icon(preview_window)

    preview_window.geometry(f"{window_width}x{window_height}")

    container = ttk.Frame(preview_window)
    container.pack(fill=tk.BOTH, expand=True)
    container.rowconfigure(0, weight=1)
    container.columnconfigure(0, weight=1)

    canvas = tk.Canvas(container, bg="white")
    v_scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=v_scrollbar.set)

    canvas.grid(row=0, column=0, sticky="nsew")
    v_scrollbar.grid(row=0, column=1, sticky="ns")

    y_offset = 20
    for photo in preview_photos:
        canvas.create_image(20, y_offset, anchor=tk.NW, image=photo)
        y_offset += photo.height() + 20

    canvas.images = preview_photos
    canvas.configure(scrollregion=canvas.bbox("all"))
    _bind_mousewheel(canvas)
    center_window(preview_window)


def extract_text(pdf_path: Path) -> str:
    with pdfplumber.open(pdf_path) as pdf:
        pages_text = [page.extract_text() or "" for page in pdf.pages]
    return "\n".join(pages_text)


def _search(pattern: str, text: str) -> str | None:
    match = re.search(pattern, text, flags=re.IGNORECASE)
    return match.group(1) if match else None


def _to_number(value: str) -> int | float | str:
    normalized = value.strip()
    if re.fullmatch(r"[+-]?\d+(?:\.\d+)?", normalized):
        return float(normalized) if "." in normalized else int(normalized)
    return value


def _extract_scan_datetime(text: str) -> tuple[str | None, str | None]:
    date_time_match = None
    for date_time_match in re.finditer(
        r"([0-9]{2}/[0-9]{2}/[0-9]{4})\s+([0-9]{2}:[0-9]{2}(?::[0-9]{2})?)",
        text,
    ):
        pass
    if date_time_match:
        return date_time_match.group(1), date_time_match.group(2)
    return None, None


def _derive_patient_id(pdf_path: Path) -> str:
    return pdf_path.stem.split("_", 1)[0]


def parse_report_text(text: str) -> dict[str, object]:
    normalized = re.sub(r"\s+", " ", text)

    patient_id = _search(r"Patient ID:\s*(\S+)", normalized)
    dob = _search(r"Date Of Birth:\s*([0-9]{2}/[0-9]{2}/[0-9]{4})", normalized)

    scan_date, scan_time = _extract_scan_datetime(normalized)

    age_gender_match = re.search(r"Age, Gender:\s*([0-9]+),\s*([A-Za-z]+)", normalized, flags=re.IGNORECASE)
    age = age_gender_match.group(1) if age_gender_match else None
    gender = age_gender_match.group(2) if age_gender_match else None

    height_cm = _search(r"Height:\s*([0-9.]+)\s*cm", normalized)
    height_m = round(float(height_cm) / 100, 2) if height_cm else None

    pulses = _search(r"Number Of Pulses:\s*([0-9]+)", normalized)

    heart_rate_period = re.search(r"Heart Rate, Period:\s*([0-9.]+)\s*bpm,\s*([0-9.]+)\s*ms", normalized, flags=re.IGNORECASE)
    heart_rate = heart_rate_period.group(1) if heart_rate_period else None
    period = heart_rate_period.group(2) if heart_rate_period else None

    ejection_match = re.search(r"Ejection Duration \(ED\):\s*([0-9.]+)\s*ms,\s*([0-9.]+)\s*%", normalized, flags=re.IGNORECASE)
    ejection_ms = ejection_match.group(1) if ejection_match else None
    ejection_pct = ejection_match.group(2) if ejection_match else None

    aortic_t2 = _search(r"Aortic T2:\s*([0-9.]+)\s*ms", normalized)
    p1_height = _search(r"P1 Height.*?:\s*([0-9.]+)\s*mmHg", normalized)
    aortic_augmentation = _search(r"Aortic Augmentation.*?:\s*([-+]?[0-9.]+)\s*mmHg", normalized)

    aix_match = re.search(r"Aortic AIx \(AP/PP, P2/P1\):\s*([-+]?[0-9.]+)\s*%,\s*([-+]?[0-9.]+)\s*%", normalized, flags=re.IGNORECASE)
    aortic_aix_ap_pp = aix_match.group(1) if aix_match else None
    aortic_aix_p2_p1 = aix_match.group(2) if aix_match else None

    aix_hr75 = _search(r"Aortic AIx \(AP/PP\) @HR75:\s*([-+]?[0-9.]+)\s*%", normalized)
    buckberg = _search(r"Buckberg SEVR:\s*([0-9.]+)\s*%", normalized)

    pti_match = re.search(r"PTI \(Systole, Diastole\):\s*([0-9.]+),\s*([0-9.]+)\s*mmHg\.s/min", normalized, flags=re.IGNORECASE)
    pti_systolic = pti_match.group(1) if pti_match else None
    pti_diastolic = pti_match.group(2) if pti_match else None

    end_systolic_pressure = _search(r"End Systolic Pressure:\s*([0-9.]+)\s*mmHg", normalized)

    map_match = re.search(r"MAP \(Systole, Diastole\):\s*([0-9.]+),\s*([0-9.]+)\s*mmHg", normalized, flags=re.IGNORECASE)
    map_systolic = map_match.group(1) if map_match else None
    map_diastolic = map_match.group(2) if map_match else None

    pulse_height = _search(r"Pulse Height:\s*([0-9.]+)", normalized)
    pulse_height_variation = _search(r"Pulse Height Variation:\s*([0-9.]+)\s*%", normalized)
    diastolic_variation = _search(r"Diastolic Variation:\s*([0-9.]+)\s*%", normalized)
    shape_deviation = _search(r"Shape Deviation:\s*([0-9.]+)\s*%", normalized)
    pulse_length_variation = _search(r"Pulse Length Variation:\s*([0-9.]+)\s*%", normalized)
    overall_quality = _search(r"Overall Quality:\s*([0-9.]+)\s*%", normalized)

    amplification = _search(r"PP Amplification:\s*([0-9.]+)\s*%", normalized)

    brachial_match = re.search(r"Brachial SYS/DIA:\s*([0-9.]+)/([0-9.]+)", normalized, flags=re.IGNORECASE)
    peripheral_sys = brachial_match.group(1) if brachial_match else None
    peripheral_dia = brachial_match.group(2) if brachial_match else None

    aortic_sys = None
    aortic_dia = None
    peripheral_pp = None
    aortic_pp = None
    peripheral_mean = None
    table_heart_rate = None

    sp_match = re.search(r"SP\s+([0-9.]+)\s+([0-9.]+)", normalized, flags=re.IGNORECASE)
    if sp_match:
        peripheral_sys = peripheral_sys or sp_match.group(1)
        aortic_sys = sp_match.group(2)

    dp_match = re.search(r"DP\s+([0-9.]+)\s+([0-9.]+)", normalized, flags=re.IGNORECASE)
    if dp_match:
        peripheral_dia = peripheral_dia or dp_match.group(1)
        aortic_dia = dp_match.group(2)

    pp_match = re.search(r"PP\s+([0-9.]+)\s+([0-9.]+)", normalized, flags=re.IGNORECASE)
    if pp_match:
        peripheral_pp = pp_match.group(1)
        aortic_pp = pp_match.group(2)

    map_hr_match = re.search(r"MAP HR\s+([0-9.]+)\s+([0-9.]+)", normalized, flags=re.IGNORECASE)
    if map_hr_match:
        peripheral_mean = map_hr_match.group(1)
        table_heart_rate = map_hr_match.group(2)

    if peripheral_sys and peripheral_dia and peripheral_pp is None:
        try:
            peripheral_pp = str(float(peripheral_sys) - float(peripheral_dia))
        except ValueError:
            peripheral_pp = None

    if aortic_sys and aortic_dia and aortic_pp is None:
        try:
            aortic_pp = str(float(aortic_sys) - float(aortic_dia))
        except ValueError:
            aortic_pp = None

    heart_rate = heart_rate or table_heart_rate

    record = {
        "Scanned ID": patient_id,
        "Scan Date": scan_date,
        "Scan Time": scan_time,
        "Date of Birth": dob,
        "Age": age,
        "Gender": gender,
        "Height (m)": height_m,
        "# of Pulses": pulses,
        "Pulse Height": pulse_height,
        "Pulse Height Variation (%)": pulse_height_variation,
        "Diastolic Variation (%)": diastolic_variation,
        "Shape Deviation (%)": shape_deviation,
        "Pulse Length Variation (%)": pulse_length_variation,
        "Overall Quality (%)": overall_quality,
        "Peripheral Systolic Pressure (mmHg)": peripheral_sys,
        "Peripheral Diastolic Pressure (mmHg)": peripheral_dia,
        "Peripheral Pulse Pressure (mmHg)": peripheral_pp,
        "Peripheral Mean Pressure (mmHg)": peripheral_mean,
        "Aortic Systolic Pressure (mmHg)": aortic_sys,
        "Aortic Diastolic Pressure (mmHg)": aortic_dia,
        "Aortic Pulse Pressure (mmHg)": aortic_pp,
        "Heart Rate (bpm)": heart_rate,
        "Pulse Pressure Amplification (%)": amplification,
        "Period (ms)": period,
        "Ejection Duration (ms)": ejection_ms,
        "Ejection Duration (%)": ejection_pct,
        "Aortic T2 (ms)": aortic_t2,
        "P1 Height (mmHg)": p1_height,
        "Aortic Augmentation (mmHg)": aortic_augmentation,
        "Aortic AIx AP/PP(%)": aortic_aix_ap_pp,
        "Aortic AIx P2/P1(%)": aortic_aix_p2_p1,
        "Aortic AIx AP/PP @ HR75 (%)": aix_hr75,
        "Buckberg SEVR (%)": buckberg,
        "PTI Systolic (mmHg.s/min)": pti_systolic,
        "PTI Diastolic (mmHg.s/min)": pti_diastolic,
        "End Systolic Pressure (mmHg)": end_systolic_pressure,
        "MAP Systolic (mmHg)": map_systolic,
        "MAP Diastolic (mmHg)": map_diastolic,
    }

    for key, value in record.items():
        if isinstance(value, str):
            record[key] = _to_number(value)
    return record


def _detect_report_type(text: str) -> str:
    normalized = text.lower()
    if DETAILED_REPORT_MARKER.lower() in normalized:
        return "detailed"
    if CLINICAL_REPORT_MARKER.lower() in normalized:
        return "clinical"
    return "unrecognized"


def _empty_record(message: str, pdf_path: Path) -> dict[str, object]:
    record: dict[str, object] = {column: None for column in COLUMNS}
    record["Source File"] = pdf_path.name
    record["Patient ID"] = message
    return record


def _prepare_dataframe(records: list[dict[str, object]]) -> tuple[pd.DataFrame, pd.Series]:
    df = pd.DataFrame(records)

    for column in COLUMNS + EXTRA_COLUMNS:
        if column not in df.columns:
            df[column] = None

    df = df[COLUMNS + EXTRA_COLUMNS]

    df["Special Row"] = df["Patient ID"].isin(
        {CLINICAL_REPORT_MESSAGE, UNRECOGNIZED_REPORT_MESSAGE}
    )

    df.loc[df["Special Row"], COLUMNS[2:]] = None

    df.sort_values(
        by=["Special Row", "Patient ID", "Scan Date", "Scan Time"], inplace=True
    )

    special_rows = df["Special Row"]
    regular_df = df.loc[~special_rows].drop_duplicates(
        subset=["Patient ID", "Scan Time", "PTI Diastolic (mmHg.s/min)"],
        keep="first",
    )
    df = pd.concat([regular_df, df.loc[special_rows]], ignore_index=True)

    df.sort_values(
        by=["Special Row", "Patient ID", "Scan Date", "Scan Time"],
        inplace=True,
        ignore_index=True,
    )

    special_row_mask = df["Special Row"].copy()

    df["Recording #"] = None
    valid_rows = ~df["Special Row"]
    df.loc[valid_rows, "Recording #"] = (
        df[valid_rows].groupby("Patient ID").cumcount() + 1
    )

    return df, special_row_mask


def process_pdf(pdf_path: Path) -> dict[str, object]:
    text = extract_text(pdf_path)
    report_type = _detect_report_type(text)

    if report_type == "detailed":
        data = parse_report_text(text)
        data["Source File"] = pdf_path.name
        data["Patient ID"] = _derive_patient_id(pdf_path)
        return data

    if report_type == "clinical":
        return _empty_record(CLINICAL_REPORT_MESSAGE, pdf_path)

    return _empty_record(UNRECOGNIZED_REPORT_MESSAGE, pdf_path)


def _closest_pair_indices(df: pd.DataFrame, fields: list[str]) -> tuple[int, int] | None:
    if len(df) < 2:
        return None

    systolic_only = fields == ["Peripheral Systolic Pressure (mmHg)"]
    diastolic_values = (
        pd.to_numeric(df["Peripheral Diastolic Pressure (mmHg)"], errors="coerce")
        if systolic_only and "Peripheral Diastolic Pressure (mmHg)" in df
        else None
    )

    min_distance = float("inf")
    min_diastolic_diff = float("inf")
    closest_pair: tuple[int, int] | None = None

    for i, idx_i in enumerate(df.index[:-1]):
        for idx_j in df.index[i + 1 :]:
            diff = df.loc[idx_i, fields] - df.loc[idx_j, fields]
            distance = (diff.pow(2).sum()) ** 0.5
            diastolic_diff = float("inf")
            if systolic_only and diastolic_values is not None:
                diastolic_diff = diastolic_values.loc[idx_i] - diastolic_values.loc[idx_j]
                diastolic_diff = abs(diastolic_diff) if pd.notna(diastolic_diff) else float("inf")

            if distance < min_distance:
                min_distance = distance
                min_diastolic_diff = diastolic_diff
                closest_pair = (idx_i, idx_j)
            elif distance == min_distance and systolic_only:
                if diastolic_diff < min_diastolic_diff:
                    min_diastolic_diff = diastolic_diff
                    closest_pair = (idx_i, idx_j)

    return closest_pair


def _average_pair_rows(pair_df: pd.DataFrame, excluded_fields: set[str]) -> dict[str, object]:
    averaged: dict[str, object] = {}
    for column in pair_df.columns:
        if column in excluded_fields:
            continue
        if column == "Patient ID":
            averaged[column] = pair_df[column].iloc[0]
            continue

        numeric_values = pd.to_numeric(pair_df[column], errors="coerce")
        if numeric_values.notna().any():
            averaged[column] = numeric_values.mean()
        else:
            non_null = pair_df[column].dropna()
            averaged[column] = non_null.iloc[0] if not non_null.empty else None

    return averaged


def _build_analyzed_data(
    df: pd.DataFrame, mode: int, manual_pairs: dict[str, tuple[int, int]] | None = None
) -> tuple[pd.DataFrame, set[int], dict[str, tuple[int, int]]]:
    analysis_fields_by_mode: dict[int, list[str]] = {
        1: [
            "Peripheral Systolic Pressure (mmHg)",
            "Peripheral Diastolic Pressure (mmHg)",
            "Peripheral Mean Pressure (mmHg)",
        ],
        2: ["Peripheral Systolic Pressure (mmHg)"],
    }

    analysis_fields = analysis_fields_by_mode.get(mode, analysis_fields_by_mode[1])

    numeric_df = df.copy()
    for field in analysis_fields:
        numeric_df[field] = pd.to_numeric(numeric_df[field], errors="coerce")

    analyzed_records: list[dict[str, object]] = []
    kept_indices: set[int] = set()
    used_pairs: dict[str, tuple[int, int]] = {}

    manual_pairs = manual_pairs or {}

    for patient_id, group in numeric_df.groupby("Patient ID"):
        valid_group = group.dropna(subset=analysis_fields)
        pair: tuple[int, int] | None = manual_pairs.get(patient_id)

        if not pair or not all(index in valid_group.index for index in pair):
            pair = _closest_pair_indices(valid_group, analysis_fields)
        if pair is None:
            continue

        pair_df = df.loc[list(pair)]
        averaged_record = _average_pair_rows(pair_df, AVERAGED_EXCLUDED_FIELDS)
        averaged_record["Patient ID"] = patient_id
        analyzed_records.append(averaged_record)
        kept_indices.update(pair)
        used_pairs[patient_id] = pair

    return pd.DataFrame(analyzed_records), kept_indices, used_pairs


def _quality_check_summary(
    df: pd.DataFrame,
    used_pairs: dict[str, tuple[int, int]],
) -> tuple[dict[str, str], list[str]]:
    regular_df = df.loc[~df["Special Row"]].copy()
    single_file_patient_ids = (
        regular_df["Patient ID"].value_counts().loc[lambda s: s == 1].index.tolist()
    )

    clinical_rows = df[df["Patient ID"] == CLINICAL_REPORT_MESSAGE]
    clinical_ids = {
        _derive_patient_id(Path(str(source)))
        for source in clinical_rows["Source File"].dropna()
    }

    diff_fields = [
        "Peripheral Systolic Pressure (mmHg)",
        "Peripheral Diastolic Pressure (mmHg)",
        "Aortic Systolic Pressure (mmHg)",
        "Aortic Diastolic Pressure (mmHg)",
    ]

    checks: dict[str, str] = {}

    for patient_id, group in regular_df.groupby("Patient ID"):
        failures: list[str] = []
        if len(group) == 1:
            failures.append("Only one file was uploaded for this participant.")

        if patient_id in clinical_ids:
            failures.append(
                "A clinical report was uploaded. Only detailed reports are used for"
                " analysis. Confirm all detailed reports are uploaded."
            )

        pair = used_pairs.get(patient_id)
        if pair:
            pair_df = df.loc[list(pair)]
            scan_dates = pd.to_datetime(
                pair_df["Scan Date"], errors="coerce", dayfirst=True
            )
            if scan_dates.notna().sum() == 2:
                unique_dates = {date.date() for date in scan_dates.dropna()}
                if len(unique_dates) > 1:
                    failures.append(
                        "Participant has different scan dates between the files choosen"
                        " for analysis."
                    )

            for field in diff_fields:
                if field not in pair_df.columns:
                    continue
                values = pd.to_numeric(pair_df[field], errors="coerce")
                if values.notna().all():
                    diff = abs(values.iloc[0] - values.iloc[1])
                    if diff > 5:
                        failures.append(
                            "The two selected files differ by more than 5 mmhg for"
                            f" {field}."
                        )

        checks[patient_id] = "Pass" if not failures else " // ".join(failures)

    return checks, single_file_patient_ids


def show_mode_choice_popup(root: tk.Misc, overview_count: int) -> bool:
    choice = {"mode": "auto"}

    window = tk.Toplevel(root)
    window.title("Analysis Mode")
    window.geometry("420x220")
    window.resizable(False, False)
    window.grab_set()
    set_app_icon(window)
    configure_popup_close(window, root)
    window.bind("<Unmap>", lambda _e: window.grab_release())
    window.bind("<Map>", lambda _e: window.grab_set())

    description = (
        f"{overview_count} records have more than 3 entries.\n\n"
        "Continue with the automatic averaging method or manually review the"
        " selections before exporting?"
    )

    ttk.Label(window, text=description, wraplength=380, justify=tk.CENTER).pack(
        pady=(20, 10), padx=15
    )

    button_frame = ttk.Frame(window)
    button_frame.pack(pady=10)

    def _select(mode: str) -> None:
        choice["mode"] = mode
        window.destroy()

    ttk.Button(button_frame, text="Use Auto Method", command=lambda: _select("auto")).pack(
        side=tk.LEFT, padx=10
    )
    ttk.Button(
        button_frame,
        text="Manual Overview",
        command=lambda: _select("manual"),
    ).pack(side=tk.LEFT, padx=10)

    window.protocol("WM_DELETE_WINDOW", lambda: terminate_application(root))
    center_window(window)
    root.wait_window(window)
    return choice["mode"] == "manual"


def _format_bp_string(sys: object, dia: object, mean: object) -> str:
    if pd.isna(sys) and pd.isna(dia) and pd.isna(mean):
        return "—"

    parts: list[str] = []
    if not pd.isna(sys) or not pd.isna(dia):
        parts.append(f"{sys or '—'}/{dia or '—'}")
    if not pd.isna(mean):
        parts.append(f"MAP {mean}")
    if not parts:
        return "—"
    if len(parts) == 1:
        return parts[0]
    return f"{parts[0]} ({parts[1]})"


class ManualOverview:
    def __init__(
        self,
        root: tk.Misc,
        df: pd.DataFrame,
        auto_pairs: dict[str, tuple[int, int]],
        manual_patients: list[str],
    ):
        self.root = root
        self.df = df
        self.auto_pairs = auto_pairs
        self.manual_patients = manual_patients
        self.current_index = 0
        self.completed = False
        self.manual_pairs: dict[str, list[int]] = {}
        self.manual_buttons: dict[int, tk.Button] = {}
        self.data_sheet_folder: Path | None = None
        self.base_font = ("TkDefaultFont", 11)
        self.value_font = ("TkDefaultFont", 11, "bold")
        self.default_button_bg = tk.Button(root).cget("bg")

        for patient_id in manual_patients:
            auto_pair = list(auto_pairs.get(patient_id, ()))
            patient_rows = self._patient_rows(patient_id)
            fallback = list(patient_rows.index[:2])
            self.manual_pairs[patient_id] = auto_pair[:2] if len(auto_pair) == 2 else fallback

        # ---- Toplevel window ----
        self.window = tk.Toplevel(root)
        self.window.title("Manual Overview")
        # Slightly wider so long filenames fit comfortably
        self.window.geometry("811x520")
        # Keep layout static – user can’t resize and stretch the grid
        self.window.resizable(False, False)
        self.window.grab_set()
        set_app_icon(self.window)
        self.window.bind("<Unmap>", self._release_grab)
        self.window.bind("<Map>", self._restore_grab)
        configure_popup_close(self.window, root)

        # row 0: header
        # row 1: content
        # row 2: bottom buttons

        # ---- Header ----
        self.header_container = ttk.Frame(self.window)
        self.header_container.grid(row=0, column=0, pady=(15, 5), padx=15, sticky="w")

        self.header_label = ttk.Label(
            self.header_container, font=("TkDefaultFont", 12, "bold")
        )
        self.header_label.pack(side=tk.LEFT)

        self.data_sheet_link = tk.Label(
            self.header_container,
            text=" [Data Collection Sheet]",
            fg="blue",
            cursor="hand2",
            font=self.base_font,
        )
        self.data_sheet_link.pack(side=tk.LEFT, padx=(10, 0))
        self.warning_icon = tk.Canvas(
            self.header_container,
            width=18,
            height=18,
            highlightthickness=0,
            cursor="hand2",
        )
        self.warning_icon.create_oval(2, 2, 16, 16, fill="#d0021b", outline="#a40015")
        self.warning_icon.create_text(
            9,
            9,
            text="!",
            fill="white",
            font=("TkDefaultFont", 11, "bold"),
        )
        self.warning_icon.pack(side=tk.LEFT, padx=(8, 0))
        self.warning_icon.bind("<Enter>", self._show_warning_tooltip)
        self.warning_icon.bind("<Leave>", self._hide_warning_tooltip)
        self.warning_tooltip: tk.Toplevel | None = None
        self.warning_messages: list[str] = []

        # ---- Content area (no scrollbars, fixed width columns) ----
        self.content_container = ttk.Frame(self.window)
        self.content_container.grid(row=1, column=0, sticky="n", padx=15, pady=(0, 5))

        self.content_frame = ttk.Frame(self.content_container)
        self.content_frame.grid(row=0, column=0, sticky="nw")

        # Fixed column widths so grid doesn’t resize with the window
        # 0: filename (made wide so full filename is visible)
        self.content_frame.columnconfigure(0, minsize=200)
        # 1–3: peripheral SYS / DIA / MAP
        self.content_frame.columnconfigure(1, minsize=80)
        self.content_frame.columnconfigure(2, minsize=80)
        self.content_frame.columnconfigure(3, minsize=80)
        # 4–5: aortic SYS / DIA
        self.content_frame.columnconfigure(4, minsize=80)
        self.content_frame.columnconfigure(5, minsize=80)
        # 6–7: Manual / Auto buttons
        self.content_frame.columnconfigure(6, minsize=90)
        self.content_frame.columnconfigure(7, minsize=90)

        # ---- Bottom controls (single bar at bottom, no right-hand pane) ----
        self.controls = ttk.Frame(self.window)
        self.controls.grid(row=2, column=0, sticky="ew", pady=(5, 15), padx=10)
        self.controls.columnconfigure(0, weight=1)
        self.controls.columnconfigure(1, weight=1)
        self.controls.columnconfigure(2, weight=1)

        self.prev_button = ttk.Button(self.controls, text="Previous", command=self._go_previous)
        self.prev_button.grid(row=0, column=0, sticky="w")

        self.save_button = ttk.Button(
            self.controls, text="Save All, Complete Analysis", command=self._complete
        )
        self.save_button.grid(row=0, column=1, sticky="ew", padx=15)

        self.next_button = ttk.Button(self.controls, text="Next", command=self._go_next)
        self.next_button.grid(row=0, column=2, sticky="e")

        self.window.protocol("WM_DELETE_WINDOW", lambda: terminate_application(root))

        self._render_patient()
        center_window(self.window)

    def _release_grab(self, _event: tk.Event | None = None) -> None:  # type: ignore[type-arg]
        self.window.grab_release()

    def _restore_grab(self, _event: tk.Event | None = None) -> None:  # type: ignore[type-arg]
        self.window.grab_set()

    def _patient_rows(self, patient_id: str) -> pd.DataFrame:
        return self.df.loc[
            (self.df["Patient ID"] == patient_id) & (self.df["Special Row"] != True)
        ]

    def _data_sheet_path(self, patient_id: str) -> Path | None:
        if self.data_sheet_folder is None or not self.data_sheet_folder.exists():
            return None

        subject_prefix = re.split(r"[ _]", patient_id, maxsplit=1)[0].lower()
        for candidate in sorted(self.data_sheet_folder.glob("*.pdf")):
            if candidate.stem.lower().startswith(subject_prefix):
                return candidate
        return None

    def _prompt_for_data_sheet_folder(self) -> bool:
        selected = filedialog.askdirectory(
            title="Select data collection sheet folder", parent=self.window
        )
        if not selected:
            return False
        self.data_sheet_folder = Path(selected)
        return True

    def _open_data_collection_sheet(self, patient_id: str) -> None:
        if self.data_sheet_folder is None or not self.data_sheet_folder.exists():
            if not self._prompt_for_data_sheet_folder():
                return

        data_sheet_path = self._data_sheet_path(patient_id)
        if data_sheet_path is None:
            retry = messagebox.askyesno(
                "Data collection sheet",
                "No matching data collection sheet was found."
                " Would you like to choose another folder?",
                parent=self.window,
            )
            if retry and self._prompt_for_data_sheet_folder():
                data_sheet_path = self._data_sheet_path(patient_id)
            else:
                return

        if data_sheet_path is None:
            messagebox.showerror(
                "Data collection sheet",
                "No matching data collection sheet was found in the selected folder.",
                parent=self.window,
            )
            return

        show_pdf_preview(self.window, data_sheet_path)

    def _update_header(self, patient_id: str) -> None:
        total = len(self.manual_patients)
        self.header_label.configure(
            text=f"Reviewing record {self.current_index + 1} of {total} — {patient_id}"
        )
        self.data_sheet_link.unbind("<Button-1>")
        self.data_sheet_link.bind(
            "<Button-1>",
            lambda _e, pid=patient_id: self._open_data_collection_sheet(pid),
        )
        self._update_warning_indicator(patient_id)

    def _patient_warnings(self, patient_id: str) -> list[str]:
        warnings: list[str] = []
        patient_rows = self._patient_rows(patient_id)
        scan_dates = pd.to_datetime(
            patient_rows["Scan Date"], errors="coerce", dayfirst=True
        )
        unique_dates = {date.date() for date in scan_dates.dropna()}
        if len(unique_dates) > 1:
            warnings.append(
                "Scan dates differ across this participant's files."
            )

        clinical_rows = self.df[self.df["Patient ID"] == CLINICAL_REPORT_MESSAGE]
        clinical_ids = {
            _derive_patient_id(Path(str(source)))
            for source in clinical_rows["Source File"].dropna()
        }
        if patient_id in clinical_ids:
            warnings.append(
                "A clinical report was uploaded. Only detailed reports are used for"
                " analysis. Confirm all detailed reports are uploaded."
            )

        return warnings

    def _update_warning_indicator(self, patient_id: str) -> None:
        self.warning_messages = self._patient_warnings(patient_id)
        if self.warning_messages:
            self._hide_warning_tooltip()
            self.warning_icon.pack_forget()
            self.warning_icon.pack(side=tk.LEFT, padx=(8, 0))
        else:
            self._hide_warning_tooltip()
            self.warning_icon.pack_forget()

    def _show_warning_tooltip(self, _event: tk.Event | None = None) -> None:  # type: ignore[type-arg]
        if not self.warning_messages or self.warning_tooltip is not None:
            return
        tooltip = tk.Toplevel(self.window)
        tooltip.wm_overrideredirect(True)
        tooltip.attributes("-topmost", True)
        label = ttk.Label(
            tooltip,
            text="\n".join(self.warning_messages),
            background="#fff4e5",
            relief="solid",
            borderwidth=1,
            font=self.base_font,
            justify=tk.LEFT,
        )
        label.pack(ipadx=6, ipady=4)
        tooltip.update_idletasks()
        window_x = self.window.winfo_rootx()
        window_y = self.window.winfo_rooty()
        window_width = self.window.winfo_width()
        window_height = self.window.winfo_height()
        tooltip_width = tooltip.winfo_reqwidth()
        tooltip_height = tooltip.winfo_reqheight()
        margin = 10
        x = self.warning_icon.winfo_rootx() + 20
        y = self.warning_icon.winfo_rooty() + 20
        x = max(window_x + margin, min(x, window_x + window_width - tooltip_width - margin))
        y = max(window_y + margin, min(y, window_y + window_height - tooltip_height - margin))
        tooltip.wm_geometry(f"+{x}+{y}")
        self.warning_tooltip = tooltip

    def _hide_warning_tooltip(self, _event: tk.Event | None = None) -> None:  # type: ignore[type-arg]
        if self.warning_tooltip is None:
            return
        self.warning_tooltip.destroy()
        self.warning_tooltip = None

    def _button_text(self, label: str, selected: bool) -> str:
        return f"{label} {CHECKMARK}" if selected else label

    def _resize_for_content(self) -> None:
        self.window.update_idletasks()
        required_height = (
            self.header_container.winfo_reqheight()
            + self.content_container.winfo_reqheight()
            + self.controls.winfo_reqheight()
            + 60
        )
        height = max(520, required_height)
        self.window.geometry(f"811x{height}")

    def _render_patient(self) -> None:
        for child in self.content_frame.winfo_children():
            child.destroy()

        patient_id = self.manual_patients[self.current_index]
        self._update_header(patient_id)
        self.manual_buttons.clear()

        patient_rows = self._patient_rows(patient_id)
        auto_pair = set(self.auto_pairs.get(patient_id, ()))
        manual_selection = self.manual_pairs.get(patient_id, [])

        ttk.Label(
            self.content_frame,
            text=(
                "Click a filename to preview the PDF. Use Manual to choose exactly two"
                " files for averaging."
            ),
            wraplength=780,
            justify=tk.LEFT,
            font=self.base_font,
        ).grid(row=0, column=0, columnspan=8, sticky="w", pady=(0, 10))

        ttk.Label(
            self.content_frame,
            text="Filename",
            font=self.value_font,
            anchor="w",
        ).grid(row=1, column=0, rowspan=2, sticky="w", padx=(0, 5))

        # Centered group header over peripheral SYS/DIA/MAP
        ttk.Label(
            self.content_frame,
            text="Peripheral",
            font=self.value_font,
            anchor="center",
        ).grid(row=1, column=1, columnspan=3, sticky="n")

        ttk.Label(
            self.content_frame, text="Systolic", font=self.base_font, anchor="center"
        ).grid(row=2, column=1, sticky="n")
        ttk.Label(
            self.content_frame, text="Diastolic", font=self.base_font, anchor="center"
        ).grid(row=2, column=2, sticky="n")
        ttk.Label(
            self.content_frame, text="MAP", font=self.base_font, anchor="center"
        ).grid(row=2, column=3, sticky="n")

        # Centered group header over aortic SYS/DIA
        ttk.Label(
            self.content_frame,
            text="Aortic",
            font=self.value_font,
            anchor="center",
        ).grid(row=1, column=4, columnspan=2, sticky="n")

        ttk.Label(
            self.content_frame, text="Systolic", font=self.base_font, anchor="center"
        ).grid(row=2, column=4, sticky="n")
        ttk.Label(
            self.content_frame, text="Diastolic", font=self.base_font, anchor="center"
        ).grid(row=2, column=5, sticky="n")

        ttk.Label(
            self.content_frame, text="Manual", font=self.value_font, anchor="center"
        ).grid(row=1, column=6, rowspan=2, sticky="n")
        ttk.Label(
            self.content_frame, text="Auto", font=self.value_font, anchor="center"
        ).grid(row=1, column=7, rowspan=2, sticky="n")

        for idx, (row_index, row) in enumerate(patient_rows.iterrows(), start=3):
            file_label = tk.Label(
                self.content_frame,
                text=row.get("Source File", ""),
                fg="blue",
                cursor="hand2",
                font=self.base_font,
                anchor="w",
                width=20,
            )
            file_label.grid(row=idx, column=0, sticky="w", padx=(0, 5), pady=4)
            source_path = row.get("Source Path")
            if source_path:
                file_label.bind(
                    "<Button-1>",
                    lambda _e, path=Path(str(source_path)): show_pdf_preview(
                        self.window, path
                    ),
                )

            sys = row.get("Peripheral Systolic Pressure (mmHg)")
            dia = row.get("Peripheral Diastolic Pressure (mmHg)")
            mean = row.get("Peripheral Mean Pressure (mmHg)")
            ttk.Label(
                self.content_frame,
                text=sys or "—",
                font=self.value_font,
                anchor="center",
            ).grid(row=idx, column=1, sticky="nsew", pady=4)
            ttk.Label(
                self.content_frame,
                text=dia or "—",
                font=self.base_font,
                anchor="center",
            ).grid(row=idx, column=2, sticky="nsew", pady=4)
            ttk.Label(
                self.content_frame,
                text=mean or "—",
                font=self.base_font,
                anchor="center",
            ).grid(row=idx, column=3, sticky="nsew", pady=4)

            a_sys = row.get("Aortic Systolic Pressure (mmHg)")
            a_dia = row.get("Aortic Diastolic Pressure (mmHg)")
            ttk.Label(
                self.content_frame,
                text=a_sys or "—",
                font=self.value_font,
                anchor="center",
            ).grid(row=idx, column=4, sticky="nsew", pady=4)
            ttk.Label(
                self.content_frame,
                text=a_dia or "—",
                font=self.base_font,
                anchor="center",
            ).grid(row=idx, column=5, sticky="nsew", pady=4)

            manual_holder = ttk.Frame(self.content_frame)
            manual_holder.grid(row=idx, column=6, sticky="e", pady=4)

            manual_selected = row_index in manual_selection
            auto_selected = row_index in auto_pair

            manual_button = tk.Button(
                manual_holder,
                text=self._button_text("Manual", manual_selected),
                command=lambda idx=row_index: self._toggle_manual(patient_id, idx),
                bg=SELECTED_COLOR if manual_selected else self.default_button_bg,
                width=9,
            )
            manual_button.pack(side=tk.LEFT, padx=2)
            self.manual_buttons[row_index] = manual_button

            tk.Button(
                self.content_frame,
                text=self._button_text("Auto", auto_selected),
                state=tk.DISABLED,
                bg=SELECTED_COLOR if auto_selected else self.default_button_bg,
                width=9,
                disabledforeground="black",
            ).grid(row=idx, column=7, sticky="e", pady=4)

        self.content_frame.update_idletasks()
        self._resize_for_content()
        self._update_nav_buttons()

    def _update_nav_buttons(self) -> None:
        self.prev_button.configure(state=tk.NORMAL if self.current_index > 0 else tk.DISABLED)
        self.next_button.configure(
            state=tk.NORMAL if self.current_index < len(self.manual_patients) - 1 else tk.DISABLED
        )

    def _toggle_manual(self, patient_id: str, row_index: int) -> None:
        selection = self.manual_pairs.setdefault(patient_id, [])
        if row_index in selection:
            selection.remove(row_index)
        elif len(selection) < 2:
            selection.append(row_index)
        else:
            messagebox.showwarning(
                "Manual Overview",
                "You can only select two files at a time for manual averaging.",
                parent=self.window,
            )
            return

        if row_index in self.manual_buttons:
            button = self.manual_buttons[row_index]
            selected = row_index in selection
            button.configure(
                text=self._button_text("Manual", selected),
                bg=SELECTED_COLOR if selected else self.default_button_bg,
            )

    def _go_previous(self) -> None:
        if self.current_index <= 0:
            return
        self.current_index -= 1
        self._render_patient()

    def _go_next(self) -> None:
        if self.current_index >= len(self.manual_patients) - 1:
            return
        self.current_index += 1
        self._render_patient()

    def _complete(self) -> None:
        for patient_id in self.manual_patients:
            selection = self.manual_pairs.get(patient_id, [])
            if len(selection) != 2:
                messagebox.showerror(
                    "Manual Overview",
                    "Please choose exactly two files for each patient before saving.",
                    parent=self.window,
                )
                return

        self.completed = True
        self.window.destroy()

    def run(self) -> dict[str, tuple[int, int]] | None:
        self.root.wait_window(self.window)
        if not self.completed:
            return None
        return {
            patient_id: tuple(selection)
            for patient_id, selection in self.manual_pairs.items()
            if len(selection) == 2
        }


def save_to_excel(
    records: list[dict[str, object]],
    output_path: Path,
    manual_pairs: dict[str, tuple[int, int]] | None = None,
) -> int:
    df, special_row_mask = _prepare_dataframe(records)

    analyzed_df, kept_indices, used_pairs = _build_analyzed_data(
        df, ANALYSIS_MODE, manual_pairs
    )

    quality_checks, single_file_patient_ids = _quality_check_summary(df, used_pairs)

    averaged_columns = [
        column
        for column in df.columns
        if column not in AVERAGED_EXCLUDED_FIELDS | {"Special Row"}
    ]

    if analyzed_df.empty:
        averaged_df = pd.DataFrame(columns=averaged_columns)
    else:
        averaged_df = analyzed_df.reindex(columns=averaged_columns).copy()

    placeholder_rows = []
    for patient_id in single_file_patient_ids:
        row = {column: "-" for column in averaged_columns}
        row["Patient ID"] = patient_id
        placeholder_rows.append(row)

    if placeholder_rows:
        averaged_df = pd.concat(
            [averaged_df, pd.DataFrame(placeholder_rows)], ignore_index=True
        )

    averaged_df["Quality Check"] = (
        averaged_df["Patient ID"].map(quality_checks).fillna("Pass")
    )

    if "Patient ID" in averaged_df.columns:
        ordered_columns = [
            column for column in averaged_df.columns if column != "Quality Check"
        ]
        insert_at = ordered_columns.index("Patient ID") + 1
        ordered_columns.insert(insert_at, "Quality Check")
        averaged_df = averaged_df[ordered_columns]
    else:
        averaged_df.insert(0, "Quality Check", averaged_df.pop("Quality Check"))

    df["Analyed"] = "No"
    if kept_indices:
        df.loc[df.index.isin(kept_indices), "Analyed"] = "Yes"

    kept_df = df[df["Analyed"] == "Yes"].copy()

    date_columns = ["Scan Date", "Date of Birth"]

    def _normalize_dates(frame: pd.DataFrame) -> pd.DataFrame:
        for date_column in date_columns:
            if date_column not in frame.columns:
                continue

            parsed_dates = pd.to_datetime(
                frame[date_column], errors="coerce", dayfirst=True
            )
            frame.loc[:, date_column] = parsed_dates
        return frame

    df = _normalize_dates(df)
    kept_df = _normalize_dates(kept_df)
    averaged_df = _normalize_dates(averaged_df)
    if single_file_patient_ids:
        placeholder_mask = averaged_df["Patient ID"].isin(single_file_patient_ids)
        fill_columns = [
            column
            for column in averaged_df.columns
            if column not in {"Patient ID", "Quality Check"}
        ]
        averaged_df.loc[placeholder_mask, fill_columns] = "-"

    def _strip_aux_columns(frame: pd.DataFrame) -> pd.DataFrame:
        return frame.drop(columns=["Special Row", *EXTRA_COLUMNS], errors="ignore")

    df_to_save = _strip_aux_columns(df.copy())
    kept_df_to_save = _strip_aux_columns(kept_df.copy())
    averaged_df_to_save = _strip_aux_columns(averaged_df.copy())

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_to_save.to_excel(writer, sheet_name="All Data", index=False)
        kept_df_to_save.to_excel(writer, sheet_name="Kept Data", index=False)
        averaged_df_to_save.to_excel(writer, sheet_name="Averaged Data", index=False)

        header_alignment = Alignment(horizontal="left")
        center_alignment = Alignment(horizontal="center")
        sheet_frames = {
            "All Data": df_to_save,
            "Kept Data": kept_df_to_save,
            "Averaged Data": averaged_df_to_save,
        }

        for sheet_name, frame in sheet_frames.items():
            sheet = writer.book[sheet_name]

            for cell in sheet[1]:
                cell.alignment = header_alignment

            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = center_alignment

            if sheet_name in {"All Data", "Kept Data"}:
                first_column = sheet.iter_cols(
                    min_col=1, max_col=1, min_row=2, max_row=sheet.max_row
                )
                for cell in next(first_column, ()):  # type: ignore[arg-type]
                    cell.alignment = header_alignment

            if sheet_name == "All Data":
                for row_index, is_special in special_row_mask.items():
                    if not is_special:
                        continue

                    patient_cell = sheet.cell(row=row_index + 2, column=2)
                    patient_cell.alignment = header_alignment

            for date_column in date_columns:
                if date_column not in frame.columns:
                    continue

                date_col_index = frame.columns.get_loc(date_column) + 1

                for cell in sheet.iter_cols(
                    min_col=date_col_index,
                    max_col=date_col_index,
                    min_row=2,
                    max_row=sheet.max_row,
                ):
                    for date_cell in cell:
                        date_cell.number_format = "MM/DD/YY"

    return len(df)


def main() -> None:
    root = tk.Tk()
    set_app_icon(root)
    root.withdraw()

    # Show startup popup with image and buttons (not wired yet)
    should_continue = show_startup_popup(root)
    if not should_continue:
        terminate_application(root)

    pdf_paths = select_input_files(root)
    if not pdf_paths:
        messagebox.showinfo("PWA Data Analyzer", "No PDF files were selected.")
        root.destroy()
        return

    output_path = select_output_file(root)
    if not output_path:
        messagebox.showinfo("PWA Data Analyzer", "No output location selected.")
        root.destroy()
        return

    loading = LoadingWindow(
        root, "Analyzing files and preparing data...", total_steps=len(pdf_paths)
    )

    records: list[dict[str, object]] = []
    for index, path in enumerate(pdf_paths, start=1):
        record = process_pdf(path)
        record["Source Path"] = str(path)
        records.append(record)
        loading.update_progress(index)

    prepared_df, _ = _prepare_dataframe(records)
    _, _, auto_pairs = _build_analyzed_data(prepared_df, ANALYSIS_MODE)

    manual_patients = [
        patient_id
        for patient_id, group in prepared_df.loc[prepared_df["Special Row"] != True]
        .groupby("Patient ID")
        if len(group) > 2
    ]

    loading.close()

    manual_pairs: dict[str, tuple[int, int]] | None = None
    if manual_patients:
        use_manual = show_mode_choice_popup(root, len(manual_patients))
        if use_manual:
            manual_overview = ManualOverview(root, prepared_df, auto_pairs, manual_patients)
            manual_pairs = manual_overview.run()
        else:
            manual_pairs = auto_pairs
    else:
        manual_pairs = auto_pairs

    if manual_pairs is None:
        manual_pairs = auto_pairs

    export_loading = LoadingWindow(root, "Creating Excel export...", total_steps=1)
    exported_count = save_to_excel(records, output_path, manual_pairs)
    export_loading.update_progress(1)
    export_loading.close()

    messagebox.showinfo(
        "PWA Data Analyzer",
        f"Exported {exported_count} record(s) to {output_path}",
    )

    root.destroy()


if __name__ == "__main__":
    main()
